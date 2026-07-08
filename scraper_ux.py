"""
Anaplan NUX Scraper
-------------------
Run the script, log in, choose a model — done.

Output: Excel file with 5 sheets:
  - All Views
  - Actions Usage Report
  - Views Usage Report
  - Modules Usage Count
  - Actions <model name>

Requirements: pip install selenium openpyxl webdriver-manager python-dotenv
              Microsoft Edge (driver is downloaded automatically)

Configuration: copy .env.example to .env and fill in your own values.
"""

import csv
import getpass
import glob
import json
import logging
import os
import re
import shutil
import sys
import tempfile
import time

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService

import models

try:
    from webdriver_manager.microsoft import EdgeChromiumDriverManager
    _WDM_AVAILABLE = True
except ImportError:
    _WDM_AVAILABLE = False

# Windows consoles/redirected output often default to cp1252, which can't
# encode the box-drawing/emoji characters printed throughout this script.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

load_dotenv()


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — Configuration via interactive prompts
# ══════════════════════════════════════════════════════════════════════════════

ANAPLAN_URLS = {
    "eu2a": "https://eu2a.app.anaplan.com/",
}

# springboard-definition-service (pages/boards) is a single global service,
# always hosted here regardless of which regional shard the model lives on.
SDS_HOST = "https://us1a.app.anaplan.com"

# ── User defaults, sourced from .env (see .env.example) ───────────────────────
# All personal/environment values live in .env, not here. Any value left unset
# in .env falls back to being prompted for interactively.
DEFAULTS = {
    "username":       os.getenv("ANAPLAN_USERNAME", ""),
    "environment":    os.getenv("ANAPLAN_ENVIRONMENT", "eu2a"),
    "use_sso":        os.getenv("ANAPLAN_USE_SSO", "false").strip().lower() in ("1", "true", "yes"),
    "output_folder":  os.getenv("ANAPLAN_OUTPUT_FOLDER", ""),
}
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_OUTPUT_FOLDER = os.path.join(os.path.expanduser("~"), "Documents", "Anaplan NUX Reports")


def _separator(char="─", width=60):
    print(char * width)


def _header(title: str):
    _separator("═")
    print(f"  {title}")
    _separator("═")


def _ask(prompt: str, default: str = "") -> str:
    """Ask the user for input, with an optional default value."""
    if default:
        answer = input(f"{prompt} [{default}]: ").strip()
        return answer if answer else default
    return input(f"{prompt}: ").strip()


def _normalise_path(p: str) -> str:
    """Convert backslashes to forward slashes so Windows paths always work."""
    return p.replace("\\", "/")


def _ask_yes_no(prompt: str, default: bool = True) -> bool:
    hint = "[Y/N, default=Y]" if default else "[Y/N, default=N]"
    answer = input(f"{prompt} {hint}: ").strip().lower()
    if not answer:
        return default
    return answer in ("y", "yes")


def _ask_choice(options: list[str], prompt: str = "Choose an option") -> int:
    """Display a numbered list and return the chosen index (0-based)."""
    for i, option in enumerate(options, 1):
        print(f"  [{i}] {option}")
    while True:
        try:
            choice = int(input(f"\n{prompt}: ")) - 1
            if 0 <= choice < len(options):
                return choice
        except ValueError:
            pass
        print("  Invalid choice, please try again.")


def _collect_config() -> dict:
    """
    Interactive wizard to collect all required configuration.
    Values in DEFAULTS are pre-filled so the user can just press Enter.
    Returns a dict with: main_url, username, password, use_basic_auth, output_folder.
    """
    _header("Anaplan NUX Scraper — Setup")
    print()
    print("This script logs into Anaplan and exports model data to Excel.")
    print("Answer the questions below. Press Enter to accept the default value.")
    print()

    # ── Environment ───────────────────────────────────────────────────────────
    _separator()
    print("STEP 1 of 4 — Anaplan environment")
    _separator()
    url_keys  = list(ANAPLAN_URLS.keys())
    url_vals  = list(ANAPLAN_URLS.values())
    default_env_idx = url_keys.index(DEFAULTS["environment"]) if DEFAULTS["environment"] in url_keys else 0
    print("\nWhich Anaplan environment do you use?")
    for i, (k, v) in enumerate(ANAPLAN_URLS.items()):
        marker = " ← default" if i == default_env_idx else ""
        print(f"  [{i+1}] {k}  →  {v}{marker}")
    while True:
        raw = input(f"\nChoose an option [default={default_env_idx+1}]: ").strip()
        if not raw:
            idx = default_env_idx
            break
        try:
            idx = int(raw) - 1
            if 0 <= idx < len(url_keys):
                break
        except ValueError:
            pass
        print("  Invalid choice, please try again.")
    main_url = url_vals[idx]
    print(f"  ✓ Environment: {main_url}\n")

    # ── Credentials ───────────────────────────────────────────────────────────
    _separator()
    print("STEP 2 of 4 — Login credentials")
    _separator()
    username = _ask("\nAnaplan email address", default=DEFAULTS["username"])
    env_password = os.getenv("ANAPLAN_PASSWORD", "")
    if env_password:
        password = env_password
        print("  ✓ Password loaded from .env")
    else:
        password = getpass.getpass("Anaplan password (hidden input): ")

    print(
        "\nDoes your organisation use SSO (Single Sign-On) to log into Anaplan?\n"
        "  Answer Y if you normally log in via Microsoft/Google/your company portal.\n"
        "  Answer N if you log in with just your Anaplan email and password."
    )
    use_sso = _ask_yes_no("Use SSO?", default=DEFAULTS["use_sso"])
    print(f"  ✓ SSO: {'yes' if use_sso else 'no'}\n")

    # ── Output folder ─────────────────────────────────────────────────────────
    _separator()
    print("STEP 3 of 4 — Output location")
    _separator()
    folder_default = _normalise_path(DEFAULTS["output_folder"]) if DEFAULTS["output_folder"] else DEFAULT_OUTPUT_FOLDER
    output_folder = _normalise_path(_ask(
        "\nFolder where the Excel file should be saved",
        default=folder_default,
    ))
    os.makedirs(output_folder, exist_ok=True)
    print(f"  ✓ Output folder: {output_folder}\n")

    # ── Summary ───────────────────────────────────────────────────────────────
    _separator()
    print("STEP 4 of 4 — Confirm")
    _separator()
    print(f"\n  Environment : {main_url}")
    print(f"  User        : {username}")
    print(f"  SSO         : {'yes' if use_sso else 'no'}")
    print(f"  Save to     : {output_folder}")
    print()

    if not _ask_yes_no("Everything correct? The script will now open the browser."):
        print("\nCancelled. Restart the script to try again.")
        sys.exit(0)

    return {
        "main_url":       main_url,
        "username":       username,
        "password":       password,
        "use_basic_auth": not use_sso,
        "output_folder":  output_folder,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  Logging
# ══════════════════════════════════════════════════════════════════════════════

def setup_logging(output_folder: str, model_name: str, timestamp: str) -> str:
    log_file = os.path.join(output_folder, f"Anaplan NUX Report - {model_name}_{timestamp}.log")
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s\t%(levelname)s\t%(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return log_file


# ══════════════════════════════════════════════════════════════════════════════
#  Helper functions
# ══════════════════════════════════════════════════════════════════════════════

def _nested(d: dict, *keys, default="") -> str:
    for key in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(key, default)
        if d is None:
            return default
    return str(d) if not isinstance(d, dict) else default


def table_to_workbook(workbook: Workbook, table: list, name: str):
    sheet = workbook.create_sheet(name, 0)
    for row in table:
        sheet.append(row)

    tbl = Table(
        displayName=re.sub(r"[^A-Za-z0-9_]", "_", name),
        ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}",
    )
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium11",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,   showColumnStripes=False,
    )
    sheet.add_table(tbl)

    for col_cells in sheet.columns:
        col_letter = get_column_letter(col_cells[0].column)
        col_width  = max(len(str(cell.value or "")) for cell in col_cells)
        if col_width > 0:
            sheet.column_dimensions[col_letter].width = min(col_width, 100)


def api_get(browser: webdriver.Remote, url: str) -> dict:
    browser.get(url)
    try:
        raw = browser.find_element(By.TAG_NAME, "pre").text
    except Exception:
        raw = browser.execute_script(
            "return document.body.innerText || document.body.textContent;"
        )
    return json.loads(raw)


# ══════════════════════════════════════════════════════════════════════════════
#  Login
# ══════════════════════════════════════════════════════════════════════════════

def _dismiss_cookie_banner(browser: webdriver.Remote):
    try:
        accept_btn = WebDriverWait(browser, 7).until(
            EC.presence_of_element_located((By.XPATH,
                "//button[normalize-space()='Accept' or "
                "normalize-space()='Accepteren' or "
                "normalize-space()='Accept All' or "
                "normalize-space()='Alles accepteren']"
            ))
        )
        browser.execute_script("arguments[0].click();", accept_btn)
        time.sleep(1)
        return
    except Exception:
        pass

    try:
        WebDriverWait(browser, 3).until(
            EC.frame_to_be_available_and_switch_to_it(
                (By.XPATH, '//iframe[@title="TrustArc Cookie Consent Manager"]')
            )
        )
        btn = WebDriverWait(browser, 3).until(
            EC.presence_of_element_located(
                (By.XPATH, "//a[contains(@class,'acceptAllButton')]")
            )
        )
        browser.execute_script("arguments[0].click();", btn)
        browser.switch_to.default_content()
        time.sleep(1)
        return
    except Exception:
        browser.switch_to.default_content()

    try:
        browser.execute_script("""
            ['[class*="cookie"]','[class*="consent"]','[class*="gdpr"]',
             '[id*="cookie"]','[id*="consent"]','[class*="privacy"]']
            .forEach(function(s) {
                document.querySelectorAll(s).forEach(function(el) {
                    el.style.setProperty('display','none','important');
                });
            });
        """)
    except Exception:
        pass


def login(browser: webdriver.Remote, config: dict):
    print("\n  → Opening browser and logging into Anaplan...")
    browser.get(config["main_url"])
    _dismiss_cookie_banner(browser)
    browser.switch_to.default_content()

    WebDriverWait(browser, 15).until(
        EC.presence_of_element_located((By.ID, "email-prelogin"))
    ).send_keys(config["username"])
    WebDriverWait(browser, 15).until(
        EC.element_to_be_clickable((By.ID, "submit-prelogin"))
    ).click()

    if config.get("use_basic_auth"):
        # Non-SSO: click the "Log in with Anaplan" button that appears after email
        WebDriverWait(browser, 15).until(
            EC.element_to_be_clickable((By.ID, "prelogin-anaplan-basic"))
        ).click()

    WebDriverWait(browser, 15).until(
        EC.presence_of_element_located((By.ID, "password"))
    ).send_keys(config["password"])
    time.sleep(1)
    WebDriverWait(browser, 15).until(
        EC.element_to_be_clickable((By.ID, "btn-login"))
    ).click()

    time.sleep(6)
    print("  ✓ Logged in.\n")


# ══════════════════════════════════════════════════════════════════════════════
#  Model selection
# ══════════════════════════════════════════════════════════════════════════════

def _select_model(browser: webdriver.Remote, config: dict) -> tuple[str, str, str, str]:
    """
    Offers configured shortcuts from models.py first, falling back to the
    live API browser if none are configured or the user wants to browse.
    Returns (model_id, model_name, workspace_guid, customer_id).
    """
    shortcuts = {
        key: m for key, m in models.MODELS.items()
        if m.get("customer_id") and m.get("workspace_id") and m.get("model_id")
    }
    if not shortcuts:
        return _choose_model_from_api(browser, config)

    _separator()
    print("MODEL SELECTION")
    _separator()
    keys = list(shortcuts.keys())
    for i, key in enumerate(keys, 1):
        print(f"  [{i}] {shortcuts[key]['name']}")
    print(f"  [{len(keys) + 1}] Browse all models…")

    while True:
        entry = input("\nChoose an option: ").strip()
        try:
            idx = int(entry) - 1
            if 0 <= idx < len(keys):
                m = shortcuts[keys[idx]]
                print(f"\n  ✓ Selected: {m['name']}")
                return m["model_id"], m["name"], m["workspace_id"], m["customer_id"]
            if idx == len(keys):
                return _choose_model_from_api(browser, config)
        except ValueError:
            pass
        print("  Invalid choice, please try again.")


def _choose_model_from_api(browser: webdriver.Remote, config: dict) -> tuple[str, str, str, str]:
    """
    Fetches all available models and lets the user choose one.
    Returns (model_id, model_name, workspace_guid, customer_id).
    """
    main_url = config["main_url"]
    print("  → Fetching available models...")

    all_models = api_get(
        browser,
        f"{main_url}/a/springboard-platform-gateway-service/models?limit=50000&offset=0",
    )["items"]

    if not all_models:
        print("\nNo models found. Check your credentials and try again.")
        sys.exit(1)

    # Group by workspace for readability
    workspaces: dict[str, list] = {}
    for m in all_models:
        ws = m.get("CurrentWorkspaceName") or m.get("WorkspaceName") or "Unknown workspace"
        workspaces.setdefault(ws, []).append(m)

    _separator()
    print(f"MODEL SELECTION — {len(all_models)} model(s) found")
    _separator()

    flat_models = []
    counter = 1
    for ws_name, models in sorted(workspaces.items()):
        print(f"\n  📁 {ws_name}")
        for m in sorted(models, key=lambda x: x["ModelName"]):
            print(f"     [{counter:3}] {m['ModelName']}")
            flat_models.append(m)
            counter += 1

    print()
    while True:
        entry = input("Type the number of the model you want to scrape: ").strip()
        try:
            idx = int(entry) - 1
            if 0 <= idx < len(flat_models):
                model = flat_models[idx]
                break
        except ValueError:
            pass
        print("  Invalid choice, please try again.")

    model_id   = model["ModelGuid"]
    model_name = model["ModelName"]
    ws_guid    = model.get("CurrentWorkspaceGuid", "")
    customer   = model.get("CustomerGuid", model.get("TenantGuid", ""))

    if not customer:
        try:
            ws_resp  = api_get(browser, f"{main_url}/1/3/workspaces/{ws_guid}")
            customer = ws_resp.get("workspace", {}).get("customerId", "")
        except Exception:
            pass

    print(f"\n  ✓ Selected: {model_name}")
    print(f"    Model ID  : {model_id}")
    print(f"    Workspace : {ws_guid}")
    print()

    return model_id, model_name, ws_guid, customer


# ══════════════════════════════════════════════════════════════════════════════
#  Download helpers
# ══════════════════════════════════════════════════════════════════════════════

def _wait_for_download(download_dir: str, before_files: set, timeout: int = 60) -> str | None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        current   = set(glob.glob(os.path.join(download_dir, "*")))
        new_files = current - before_files
        complete  = [f for f in new_files if not f.endswith((".crdownload", ".tmp", ".part"))]
        if complete:
            time.sleep(0.5)
            return max(complete, key=os.path.getmtime)
        time.sleep(0.5)
    return None


def _parse_downloaded_file(filepath: str) -> tuple[list, list]:
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook as _lw
        wb = _lw(filepath, read_only=True, data_only=True)
        ws = wb.active
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        if not all_rows:
            return [], []
        return [str(h) if h is not None else "" for h in all_rows[0]], all_rows[1:]
    else:
        with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            all_rows = list(reader)
        if not all_rows:
            return [], []
        return all_rows[0], all_rows[1:]


def _click_export_button(browser: webdriver.Remote, timeout: int = 15) -> bool:
    selectors = [
        "//button[normalize-space()='Export']",
        "//a[normalize-space()='Export']",
        "//button[contains(normalize-space(.), 'Export')]",
        "//*[@role='button'][contains(normalize-space(.), 'Export')]",
        "//button[@aria-label='Export']",
        "//button[@title='Export']",
        "//*[@data-action='export']",
        "//*[@data-testid='export-button']",
    ]
    for sel in selectors:
        try:
            btn = WebDriverWait(browser, timeout).until(
                EC.element_to_be_clickable((By.XPATH, sel))
            )
            browser.execute_script("arguments[0].scrollIntoView(true);", btn)
            time.sleep(0.3)
            browser.execute_script("arguments[0].click();", btn)
            return True
        except Exception:
            continue
    return False


def _confirm_export_dialog(browser: webdriver.Remote):
    try:
        btn = WebDriverWait(browser, 6).until(
            EC.element_to_be_clickable((By.XPATH,
                "//button[normalize-space()='OK' or "
                "normalize-space()='Download' or "
                "normalize-space()='Export' or "
                "normalize-space()='CSV' or "
                "normalize-space()='Bevestigen']"
            ))
        )
        browser.execute_script("arguments[0].click();", btn)
        time.sleep(1)
    except Exception:
        pass


def _try_modeling_section_export(
    browser: webdriver.Remote,
    model_new_url: str,
    section: str,
    download_dir: str,
    extra_wait: int = 5,
    download_timeout: int = 60,
) -> tuple[list, list]:
    url_candidates = [
        f"{model_new_url}/model-settings/{section}",
        f"{model_new_url}/{section}",
        f"{model_new_url}/settings/{section}",
        f"{model_new_url}/model-settings",
    ]

    before_files = set(glob.glob(os.path.join(download_dir, "*")))
    page_loaded  = False

    for url in url_candidates:
        try:
            browser.get(url)
            WebDriverWait(browser, 8).until(
                EC.any_of(
                    EC.presence_of_element_located((By.XPATH,
                        "//button[contains(normalize-space(.), 'Export')]")),
                    EC.presence_of_element_located((By.TAG_NAME, "table")),
                    EC.presence_of_element_located((By.XPATH,
                        "//*[contains(@class,'ag-root')]")),
                    EC.presence_of_element_located((By.XPATH,
                        "//*[contains(@class,'virtualized')]")),
                )
            )
            page_loaded = True
            break
        except Exception:
            continue

    if not page_loaded:
        return [], []

    time.sleep(min(extra_wait, 4))

    if not _click_export_button(browser, timeout=8):
        return [], []

    _confirm_export_dialog(browser)

    downloaded = _wait_for_download(download_dir, before_files, timeout=download_timeout)
    if not downloaded:
        return [], []

    try:
        return _parse_downloaded_file(downloaded)
    except Exception:
        return [], []


# ══════════════════════════════════════════════════════════════════════════════
#  Actions detail builders
# ══════════════════════════════════════════════════════════════════════════════

_TGT_TYPE_MAP = {
    "MODULE_DATA":      "MODULE",
    "VERSIONS":         "VERSIONS",
    "LIST_ACCESS_DATA": "LIST",
    "USERS":            "USERS",
    "ROLE_ACCESS_DATA": "ROLES",
}


def _build_actions_detail_fallback(all_actions_raw: list) -> list:
    table = [[
        "Action Name", "Action Type",
        "Source", "Source Object", "Source Type",
        "Target Object", "Target Type",
        "Production Data",
    ]]
    for action in all_actions_raw:
        atype    = action.get("__atype", "")
        name     = action.get("name", "")
        imp_type = action.get("importType", "")
        ds_id    = action.get("importDataSourceId", "")

        if atype == "imports":
            src_type = "FILE" if ds_id else ("-" if imp_type == "VERSIONS" else "SAVED VIEW")
            tgt_type = _TGT_TYPE_MAP.get(imp_type, imp_type)
        elif atype == "exports":
            src_type = "SAVED VIEW"
            tgt_type = "FILE"
        else:
            src_type = tgt_type = ""

        table.append([name, atype, "", "", src_type, "", tgt_type, ""])
    return table


def _build_actions_detail_from_download(headers: list, rows: list) -> list:
    def _find(row_dict, *candidates):
        for c in candidates:
            v = row_dict.get(c)
            if v not in (None, ""):
                return str(v)
        return ""

    table = [[
        "Action Name", "Action Type",
        "Source", "Source Object", "Source Type",
        "Target Object", "Target Type",
        "Production Data",
    ]]
    for row_list in rows:
        if not any(v for v in row_list):
            continue
        row = dict(zip(headers, row_list))
        table.append([
            _find(row, "Name", "Action Name", "Naam", "Actienaam"),
            _find(row, "Type", "Action Type", "Type actie"),
            _find(row, "Source", "Bron", "Source Model", "Bronmodel"),
            _find(row, "Source Object", "Bronobject", "Source View", "Bronweergave"),
            _find(row, "Source Type", "Brontype", "Import Type", "Importtype"),
            _find(row, "Target", "Target Object", "Doelobject", "Target Module"),
            _find(row, "Target Type", "Doeltype"),
            _find(row, "Production Data", "Productiedata", "Production"),
        ])
    return table


# ══════════════════════════════════════════════════════════════════════════════
#  Scrapen
# ══════════════════════════════════════════════════════════════════════════════

def _progress(current: int, total: int, label: str = ""):
    pct   = int(current / total * 100) if total else 0
    bar   = "█" * (pct // 5) + "░" * (20 - pct // 5)
    print(f"\r  [{bar}] {pct:3}%  {current}/{total}  {label:<40}", end="", flush=True)


def scrape(browser: webdriver.Remote, config: dict, model_id: str, model_name: str,
           ws_guid: str, customer: str, download_dir: str):

    main_url  = config["main_url"]
    out_dir   = config["output_folder"]
    sds_url   = f"{SDS_HOST}/a/springboard-definition-service"

    model_api_url   = f"{main_url}/2/0/models/{model_id}"
    model_pages_url = f"{sds_url}/customer/{customer}/model/{model_id}/pages"
    model_new_url   = (
        f"{main_url}/a/modeling/customers/{customer}"
        f"/workspaces/{ws_guid}/models/{model_id}"
    )

    nux_re    = re.compile(r"^[0-9a-fA-F]{8}-(?:[0-9a-fA-F]{4}-){3}[0-9a-fA-F]{12}$")
    timestamp = time.strftime("%Y%m%d_%H%M%S")

    setup_logging(out_dir, model_name, timestamp)
    logging.info(f"Model: {model_name}  |  ID: {model_id}")

    workbook    = Workbook()
    blank_sheet = workbook.active

    # ── All Views ──────────────────────────────────────────────────────────────
    print("\n  → Step 1/5: Fetching views...")
    raw_views = api_get(browser, f"{model_api_url}/views")["views"]
    table = [["Name", "ID", "Module ID"]]
    for v in raw_views:
        table.append([v["name"], v["id"], v["moduleId"]])
    table_to_workbook(workbook, table, "All Views")
    workbook.remove(blank_sheet)
    views = {v["id"]: {"name": v["name"], "module": v["moduleId"]} for v in raw_views}
    print(f"  ✓ {len(raw_views)} views found.")

    # ── Modules ────────────────────────────────────────────────────────────────
    print("  → Step 2/5: Fetching modules...")
    raw_modules = api_get(browser, f"{model_api_url}/modules")["modules"]
    modules = {m["id"]: {"name": m["name"], "count": 0} for m in raw_modules}
    print(f"  ✓ {len(raw_modules)} modules found.")

    # ── Pages ─────────────────────────────────────────────────────────────────
    print("  → Step 3/5: Processing pages, actions and views...")
    pages = api_get(browser, model_pages_url)["items"]

    # ── Action definitions ─────────────────────────────────────────────────────
    actions_dict: dict = {}
    all_actions_raw: list = []
    for atype in ["processes", "imports", "exports", "actions"]:
        resp = api_get(browser, f"{model_api_url}/{atype}")
        for action in resp.get(atype, []):
            actions_dict[action["id"]] = {"name": action["name"], "type": atype}
            all_actions_raw.append({"__atype": atype, **action})

    # ── Actions Usage Report ───────────────────────────────────────────────────
    actions_table = [[
        "Action name", "Action type", "App name", "Page name",
        "Page URL", "Action ID", "App ID", "Page ID",
    ]]

    views_table = [[
        "Module/View name", "App name", "Page name",
        "View URL", "Page URL", "Module/View ID", "App ID", "Page ID",
    ]]

    for i, page in enumerate(pages):
        _progress(i + 1, len(pages), page["name"][:40])

        # Actions per pagina
        if page["pageType"] in ("BOARD", "GRID-PAGE"):
            try:
                page_content = api_get(
                    browser, f"{sds_url}/{page['pageType'].lower()}s/{page['guid']}"
                )
            except Exception:
                page_content = {}

            page_actions: set = set()
            if page["pageType"] == "BOARD":
                widgets = []
                for row in page_content.get("rows", []):
                    for col in row.get("columns", []):
                        widgets += col.get("widgets", [])
            else:
                widgets = page_content.get("widgets", [])
                page_actions.update(
                    a["actionId"] for a in page_content.get("actions", [])
                )

            for widget in widgets:
                wdef = widget.get("widgetDefinition", {})
                for a in wdef.get("widgetActions", []):
                    page_actions.add(a["actionId"])
                if wdef.get("type") == "ACTION":
                    for a in json.loads(wdef.get("actions", "[]")):
                        page_actions.add(a["id"])

            for aid in page_actions:
                if aid not in actions_dict:
                    continue
                actions_table.append([
                    actions_dict[aid]["name"],
                    actions_dict[aid]["type"],
                    page["appName"],
                    page["name"],
                    f"{main_url}/a/apps/app/{page['appGuid']}/boards/{page['guid']}",
                    aid,
                    page["appGuid"],
                    page["guid"],
                ])

        # Views per pagina
        try:
            page_data = api_get(
                browser, f"{model_pages_url}/{page['guid']}?moduleUsage=true"
            )
        except Exception:
            page_data = {}

        data_sources = {
            card["dataSourceId"]
            for card in page_data.get("pageDataSources", [])
            if card.get("dataSourceId")
        }
        for ds_id in data_sources:
            if ds_id not in views:
                continue
            views_table.append([
                views[ds_id]["name"],
                page["appName"],
                page["name"],
                f"{model_new_url}/tabs/{ds_id}",
                f"{main_url}/a/apps/app/{page['appGuid']}/boards/{page['guid']}",
                ds_id,
                page["appGuid"],
                page["guid"],
            ])
            modules[views[ds_id]["module"]]["count"] += 1

    print()  # newline after progress bar
    table_to_workbook(workbook, actions_table, "Actions Usage Report")
    table_to_workbook(workbook, views_table,   "Views Usage Report")

    # ── Modules Usage Count ────────────────────────────────────────────────────
    modules_table = [["Name", "ID", "Count"]]
    for mid, mdata in modules.items():
        modules_table.append([mdata["name"], mid, mdata["count"]])
    table_to_workbook(workbook, modules_table, "Modules Usage Count")

    # ── Actions detail via Modeling UI Export ──────────────────────────────────
    print("  → Step 4/5: Downloading actions detail via Modeling UI...")
    sheet_label     = f"Actions {model_name}"[:31]
    act_detail_table = None

    for section in ("imports", "actions"):
        headers, rows = _try_modeling_section_export(
            browser, model_new_url, section, download_dir, extra_wait=4, download_timeout=20,
        )
        if headers and rows:
            act_detail_table = _build_actions_detail_from_download(headers, rows)
            break

    if act_detail_table is None:
        logging.warning("UI export failed. Falling back to REST API data.")
        act_detail_table = _build_actions_detail_fallback(all_actions_raw)

    table_to_workbook(workbook, act_detail_table, sheet_label)

    # ── Save ───────────────────────────────────────────────────────────────────
    print("  → Step 5/5: Saving Excel file...")
    excel_path = "/".join([out_dir.rstrip("/"), f"Anaplan NUX Report - {model_name}_{timestamp}.xlsx"])
    workbook.save(excel_path)

    _separator("═")
    print(f"\n  ✅ DONE!\n")
    print(f"  File saved to:\n  {excel_path}")
    print(f"\n  Summary:")
    print(f"    • {len(raw_views)} views")
    print(f"    • {len(actions_table) - 1} action links on pages")
    print(f"    • {len(views_table) - 1} view links on pages")
    print(f"    • {len(raw_modules)} modules")
    _separator("═")


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def _get_edge_version() -> str:
    """Read the installed Edge version from the Windows registry."""
    try:
        import winreg
        for key_path in (
            r"SOFTWARE\Microsoft\Edge\BLBeacon",
            r"SOFTWARE\WOW6432Node\Microsoft\Edge\BLBeacon",
        ):
            try:
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path)
                version, _ = winreg.QueryValueEx(key, "version")
                winreg.CloseKey(key)
                return version
            except Exception:
                continue
    except Exception:
        pass
    return ""


def _create_browser(download_dir: str) -> webdriver.Remote:
    """
    Start Edge with the correct WebDriver.

    Strategy (in order):
    1. msedgedriver.exe in the same folder as this script
    2. Selenium Manager (bundled with Selenium 4.6+)
    3. webdriver-manager (needs internet on first run)
    """
    options = webdriver.EdgeOptions()
    options.add_argument("--disable-extensions")
    options.add_experimental_option("prefs", {
        "download.default_directory":   download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade":   True,
        "safebrowsing.enabled":         True,
    })

    # Strategy 1: driver next to this script
    script_dir   = os.path.dirname(os.path.abspath(__file__))
    local_driver = os.path.join(script_dir, "msedgedriver.exe")
    if os.path.isfile(local_driver):
        print(f"  → Using local driver: {local_driver}")
        try:
            service = EdgeService(executable_path=local_driver)
            browser = webdriver.Edge(service=service, options=options)
            print("  ✓ Edge driver ready.\n")
            return browser
        except Exception as e:
            print(f"  ! Local driver failed: {e}")

    # Strategy 2: Selenium Manager
    print("  → Trying Selenium Manager...")
    try:
        browser = webdriver.Edge(options=options)
        print("  ✓ Edge driver ready (Selenium Manager).\n")
        return browser
    except Exception as e:
        print(f"  ! Selenium Manager failed: {e}")

    # Strategy 3: webdriver-manager (needs internet)
    if _WDM_AVAILABLE:
        print("  → Trying webdriver-manager (needs internet)...")
        try:
            service = EdgeService(EdgeChromiumDriverManager().install())
            browser = webdriver.Edge(service=service, options=options)
            print("  ✓ Edge driver ready (webdriver-manager).\n")
            return browser
        except Exception as e:
            print(f"  ! webdriver-manager failed: {e}")

    raise RuntimeError(
        f"No working msedgedriver found.\n"
        f"Place msedgedriver.exe (matching your Edge version) in:\n  {script_dir}\n"
        f"Download from: https://developer.microsoft.com/en-us/microsoft-edge/tools/webdriver/"
    )


def main():
    # 1. Collect configuration via wizard
    config = _collect_config()

    download_dir = tempfile.mkdtemp(prefix="anaplan_dl_")

    # 2. Start browser
    print("\n  → Starting Edge browser...")
    try:
        browser = _create_browser(download_dir)
    except Exception as e:
        print(f"\n  ERROR: Could not start Edge browser.\n  {e}")
        print(
            "\n  Make sure Microsoft Edge is installed and that the Edge WebDriver\n"
            "  version matches your Edge version.\n"
            "  Run:  pip install webdriver-manager\n"
            "  to let the script manage this automatically."
        )
        shutil.rmtree(download_dir, ignore_errors=True)
        sys.exit(1)

    browser.set_script_timeout(120)

    try:
        # 3. Login
        login(browser, config)

        # 4. Model selection
        model_id, model_name, ws_guid, customer = _select_model(browser, config)

        # 5. Confirm before scraping
        if not _ask_yes_no(f"Ready to scrape '{model_name}'. Continue?"):
            print("Cancelled.")
            return

        # 6. Scrape
        scrape(browser, config, model_id, model_name, ws_guid, customer, download_dir)

        # 7. Scrape another model?
        while _ask_yes_no("\nWould you like to scrape another model?"):
            model_id, model_name, ws_guid, customer = _select_model(browser, config)
            if _ask_yes_no(f"Scrape '{model_name}'?"):
                scrape(browser, config, model_id, model_name, ws_guid, customer, download_dir)

    except KeyboardInterrupt:
        print("\n\nCancelled by user.")
    finally:
        browser.quit()
        shutil.rmtree(download_dir, ignore_errors=True)

    print("\nBrowser closed. Goodbye!")


if __name__ == "__main__":
    main()