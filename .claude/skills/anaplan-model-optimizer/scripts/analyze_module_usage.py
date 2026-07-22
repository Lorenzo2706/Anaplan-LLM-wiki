"""
Cross-reference a scraper_ux.py NUX report against a model's raw CSV export
to find modules - and, for modules that survive that check, individual line
items - that are genuinely unused, as opposed to ones that are merely absent
from the new-UX pages/boards by design (Data/Load/Calculation modules in the
DISCO pattern are normal to have zero NUX exposure - they feed other modules
via formulas instead). Classic dashboards are legacy and are not considered -
only NUX usage counts as front-end exposure.

Pass 1 (modules): a module is only reported as a deletion candidate when ALL
of these are true:
  - zero NUX usage (Modules Usage Count sheet == 0)
  - not referenced by any other module's formula (Modules.csv "Referenced By")
  - none of its line items are used as a NUX page filter (Excel "UI Filters"
    sheet)
  - not the source/target of an import or export (Imports.csv, and the
    Excel's per-model Actions detail sheet)
  - not a module category header row (e.g. "◼️ LOAD MODULES") - these are
    always fully empty and exist purely for audit-trail grouping, so they are
    reported as kept rather than filtered out silently

Pass 2 (line items): scoped to modules whose Pass-1 verdict was ACTIVE or
KEEP. A line item is a deletion candidate when it has no NUX front-end
exposure (Views Usage Report - Line Items / UI Filters sheets), no formula
reference (Line Items.csv "Referenced By"), no best-effort import/export
dot-notation match, and isn't a section-header/divider name (e.g.
"---Technical---") - overlaid with the same manual-deletion-marker annotation
(Notes/Functional Area) used in Pass 1. Line items named like a conditional-
formatting driver (e.g. "CF Overwrite", "CF - Input") get a distinct "user to
verify" verdict instead of a candidate verdict, since CF usage inside a
module's formatting rules isn't visible anywhere in the NUX scrape.

Usage:
    python analyze_module_usage.py --excel "<path to NUX report.xlsx>" \
        --model-dir "raw/models/<Model Name>" \
        [--out-json report.json] [--out-markdown report.md]

Always prints a markdown report to stdout.
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

FIXED_SHEETS = {"All Views", "Actions Usage Report", "Views Usage Report", "Modules Usage Count",
                 "Views Usage Report - Line Items", "UI Filters"}
SECTION_HEADER_PREFIX = "◼"  # "◼️" - pseudo-header rows like "◼️ LOAD MODULES"

# Manual deletion markers: model-owner intent captured in free-text Notes, or
# (at module level) a Functional Area that itself reads "DELETE". Detection
# here is purely additive - it must never suppress the reference/front-end
# safety check, only annotate the result (see analyze() / analyze_line_items()).
DELETE_MARKER_KEYWORDS = ["delete", "to be deleted", "obsolete", "deprecated", "remove"]


def detect_manual_marker(notes: str, functional_area: str = "") -> dict:
    """Scan Notes/Functional Area for a model-owner deletion marker.

    Returns {"flagged": bool, "reasons": [str, ...]}. Never returns more than
    one Notes-derived reason and one Functional-Area-derived reason, even if
    several keywords independently match the same Notes text.
    """
    notes_l = (notes or "").lower()
    matched_keywords = [kw for kw in DELETE_MARKER_KEYWORDS if kw in notes_l]
    reasons = []
    if matched_keywords:
        reasons.append("Notes: " + ", ".join(f"'{kw}'" for kw in matched_keywords))
    if "delete" in (functional_area or "").lower():
        reasons.append("Functional Area contains 'DELETE'")
    return {"flagged": bool(reasons), "reasons": reasons}


def _read_csv_rows(path: Path, delimiter_candidates=(",", ";")):
    if not path.is_file():
        return []
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(raw.splitlines()[0], delimiters="".join(delimiter_candidates))
        delimiter = dialect.delimiter
    except Exception:
        delimiter = delimiter_candidates[0]
    reader = csv.DictReader(raw.splitlines(), delimiter=delimiter)
    return list(reader)


def _extract_names(field: str) -> list:
    if not field or not field.strip():
        return []
    quoted = re.findall(r"'([^']+)'", field)
    if quoted:
        return [q.strip() for q in quoted]
    return [p.strip() for p in field.split(",") if p.strip()]


def load_modules_csv(model_dir: Path) -> dict:
    rows = _read_csv_rows(model_dir / "Modules.csv")
    modules = {}
    for row in rows:
        name = (row.get("") or "").strip()
        # Anaplan pads section-header rows (e.g. "◼️ CALCULATION MODULES") with
        # leading U+2800 (braille pattern blank) characters, so a plain
        # startswith() on the glyph misses them - strip that padding first.
        stripped = name.lstrip("⠀ \t")
        if not stripped:
            continue
        functional_area = (row.get("Functional Area") or "").strip()
        notes = (row.get("Notes") or "").strip()
        modules[name] = {
            "functional_area": functional_area,
            "referenced_by": _extract_names(row.get("Referenced By", "")),
            "is_category_header": stripped.startswith(SECTION_HEADER_PREFIX),
            "notes": notes,
            "manual_marker": detect_manual_marker(notes, functional_area),
        }
    return modules


def load_imports_csv(model_dir: Path) -> set:
    rows = _read_csv_rows(model_dir / "Imports.csv")
    used = set()
    for row in rows:
        if (row.get("Target Type") or "").strip().upper() == "MODULE":
            tgt = (row.get("Target Object") or "").strip()
            if tgt:
                used.add(tgt)
        if (row.get("Source Type") or "").strip().upper() == "MODULE":
            src = (row.get("Source Object") or "").strip()
            if src:
                used.add(src)
    return used


def load_line_items_csv(model_dir: Path) -> dict:
    """Load Line Items.csv keyed by (module_name, line_item_name).

    Line Items.csv carries its own per-line-item "Referenced By" and "Notes"
    columns - unlike Modules.csv's aggregate signals, these are already at
    the exact grain Pass 2 needs, no extraction/aggregation required.
    """
    rows = _read_csv_rows(model_dir / "Line Items.csv")
    items = {}
    for row in rows:
        name = (row.get("") or "").strip()
        module_name = (row.get("Module Name") or "").strip()
        if not name or not module_name:
            continue
        items[(module_name, name)] = {
            "notes": (row.get("Notes") or "").strip(),
            "referenced_by": _extract_names(row.get("Referenced By", "")),
        }
    return items


_DOTTED_REFERENCE_RE = re.compile(r"'([^']+)'\.(.+)$")


def _parse_dotted_reference(obj: str):
    """Extract (module_name, line_item_name) from an Imports.csv Source/Target
    Object string's final path segment, e.g.
    "Data Hub 2.0 / 'SYS 05. Time Settings'.Current Period" ->
    ("SYS 05. Time Settings", "Current Period"). Module names routinely
    contain periods (e.g. "SYS 05."), which is why only the quoted-module
    pattern is trusted - a whole-module reference like "SM 02. General
    Settings" has no quotes and correctly returns None.
    """
    if not obj:
        return None
    last_segment = obj.split(" / ")[-1].strip()
    match = _DOTTED_REFERENCE_RE.search(last_segment)
    if not match:
        return None
    return match.group(1).strip(), match.group(2).strip()


def load_import_line_item_matches(model_dir: Path, known_pairs: set) -> set:
    """Best-effort: which (module, line item) pairs already known from
    Line Items.csv are named via dot-notation in Imports.csv Source/Target
    Object. Only ever returns a subset of known_pairs - never invents a pair
    that isn't already in this model's own Line Items.csv (which would
    happen if a dot-notation reference happened to resolve to a different
    model's module/line item of the same name).
    """
    rows = _read_csv_rows(model_dir / "Imports.csv")
    matched = set()
    for row in rows:
        for field in ("Source Object", "Target Object"):
            parsed = _parse_dotted_reference((row.get(field) or "").strip())
            if parsed and parsed in known_pairs:
                matched.add(parsed)
    return matched


def load_excel(excel_path: Path):
    wb = load_workbook(excel_path, data_only=True, read_only=True)

    ux_counts = {}
    if "Modules Usage Count" in wb.sheetnames:
        ws = wb["Modules Usage Count"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            name, _id, count = (row + (None, None, None))[:3]
            ux_counts[str(name).strip()] = int(count) if count is not None else 0

    action_usage = set()
    action_sheet_names = [s for s in wb.sheetnames if s not in FIXED_SHEETS]
    if len(action_sheet_names) == 1:
        ws = wb[action_sheet_names[0]]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            idx = {h: i for i, h in enumerate(header)}
            for row in rows[1:]:
                if not row or not any(row):
                    continue
                if idx.get("Target Type") is not None and str(row[idx["Target Type"]] or "").strip().upper() == "MODULE":
                    tgt = str(row[idx["Target Object"]] or "").strip()
                    if tgt:
                        action_usage.add(tgt)
                if idx.get("Source Type") is not None and str(row[idx["Source Type"]] or "").strip().upper() == "MODULE":
                    src = str(row[idx["Source Object"]] or "").strip()
                    if src:
                        action_usage.add(src)
    elif len(action_sheet_names) > 1:
        print(
            f"WARNING: expected exactly one non-fixed sheet for the per-model actions "
            f"detail, found {action_sheet_names}. Skipping the Excel-based action-usage signal "
            f"(Imports.csv is still used).",
            file=sys.stderr,
        )

    line_item_exposure = set()
    if "Views Usage Report - Line Items" in wb.sheetnames:
        ws = wb["Views Usage Report - Line Items"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            idx = {h: i for i, h in enumerate(header)}
            mod_col, li_col = idx.get("Module/View name"), idx.get("Line Item")
            if mod_col is not None and li_col is not None:
                for row in rows[1:]:
                    if not row:
                        continue
                    module_name = str(row[mod_col] or "").strip()
                    li_name = str(row[li_col] or "").strip()
                    if module_name and li_name:
                        line_item_exposure.add((module_name, li_name))

    ui_filter_modules = set()
    if "UI Filters" in wb.sheetnames:
        ws = wb["UI Filters"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            idx = {h: i for i, h in enumerate(header)}
            mod_col = idx.get("Module")
            filter_cols = [idx[c] for c in ("Filter Column", "Filter Rows") if idx.get(c) is not None]
            if mod_col is not None:
                for row in rows[1:]:
                    if not row:
                        continue
                    module_name = str(row[mod_col] or "").strip()
                    if not module_name:
                        continue
                    for col_idx in filter_cols:
                        raw = str(row[col_idx] or "").strip()
                        for li_name in [x.strip() for x in raw.split(";") if x.strip()]:
                            line_item_exposure.add((module_name, li_name))
                            ui_filter_modules.add(module_name)

    wb.close()
    return ux_counts, action_usage, line_item_exposure, ui_filter_modules


LI_CANDIDATE_VERDICT = "CANDIDATE FOR REVIEW - no front-end exposure, no formula reference, no import/export match found"
LI_VERIFY_VERDICT = "USER TO VERIFY - conditional formatting driver, not visible in the NUX scrape"

# Section-header/divider rows kept for audit-trail readability, e.g.
# "-- Technical --", "---CF---", "== NEW ==" - decorated on both ends with
# dashes/equals, never themselves real calculation line items.
_HEADER_DIVIDER_RE = re.compile(r"^[-=]{1,}.*[-=]{1,}$")
# Conditional-formatting driver line items, e.g. "CF Overwrite", "CF - Input".
_CF_PREFIX_RE = re.compile(r"^CF(\s|-|$)")


def _is_line_item_header(name: str) -> bool:
    stripped = name.strip()
    return bool(stripped) and bool(_HEADER_DIVIDER_RE.match(stripped)) and any(c.isalnum() for c in stripped)


def _is_cf_conditional_formatting(name: str) -> bool:
    return bool(_CF_PREFIX_RE.match(name.strip()))


def _li_sort_rank(verdict: str) -> int:
    if verdict == LI_CANDIDATE_VERDICT:
        return 0
    if verdict == LI_VERIFY_VERDICT:
        return 1
    return 2


def analyze_line_items(model_dir: Path, module_results: list, line_item_exposure: set) -> dict:
    """Pass 2: per-line-item verdicts, scoped to modules whose Pass-1 verdict
    is ACTIVE or KEEP. Line items belonging to a Pass-1 CANDIDATE or DATA
    MISMATCH module are excluded entirely - deleting the module makes them
    moot, and listing them separately would just be noise.
    """
    in_scope_modules = {
        e["name"]: e for e in module_results
        if not e["verdict"].startswith("CANDIDATE") and not e["verdict"].startswith("DATA MISMATCH")
    }

    line_items = load_line_items_csv(model_dir)
    known_pairs = set(line_items.keys())
    import_matches = load_import_line_item_matches(model_dir, known_pairs)

    # Data-quality note: a Module/View name that appears in the NUX report's
    # front-end exposure sheets (Views Usage Report - Line Items / UI
    # Filters) but doesn't match any known module name in this model's own
    # CSV export - typically because the view was renamed since the last
    # scrape. Flagging this separately means we neither silently drop that
    # row's exposure signal nor misreport it as a mismatch on every line item
    # inside it (there is no line-item-level DATA MISMATCH verdict - see the
    # design spec).
    known_module_names = {e["name"] for e in module_results}
    exposure_module_names = {name for (name, _li) in line_item_exposure}
    unresolved_view_names = sorted(exposure_module_names - known_module_names)

    results = []
    for (module_name, li_name), meta in line_items.items():
        module_entry = in_scope_modules.get(module_name)
        if module_entry is None:
            continue

        own_marker = detect_manual_marker(meta["notes"])
        inherited_delete_flag = "delete" in (module_entry["functional_area"] or "").lower()

        exposed = (module_name, li_name) in line_item_exposure
        referenced = bool(meta["referenced_by"])
        imported = (module_name, li_name) in import_matches

        if exposed:
            verdict = "ACTIVE - shown or filtered on in the front end"
        elif referenced:
            verdict = "KEEP - feeds other line items via formula"
        elif imported:
            verdict = "KEEP - used as an import/export source or target"
        elif _is_line_item_header(li_name):
            verdict = "KEEP - section header/divider (audit trail)"
        elif _is_cf_conditional_formatting(li_name):
            verdict = LI_VERIFY_VERDICT
        else:
            verdict = LI_CANDIDATE_VERDICT

        results.append({
            "module": module_name,
            "line_item": li_name,
            "functional_area": module_entry["functional_area"],
            "verdict": verdict,
            "referenced_by": meta["referenced_by"],
            "manual_marker": own_marker,
            "inherited_module_delete_flag": inherited_delete_flag,
            "flagged_but_kept": (own_marker["flagged"] or inherited_delete_flag) and verdict.startswith(("ACTIVE", "KEEP")),
        })

    results.sort(key=lambda e: (
        e["module"],
        _li_sort_rank(e["verdict"]),
        not (e["manual_marker"]["flagged"] or e["inherited_module_delete_flag"]),
        e["line_item"],
    ))

    by_module = {}
    for e in results:
        slot = by_module.setdefault(e["module"], {
            "functional_area": e["functional_area"], "candidates": [], "to_verify": [], "flagged_but_kept": [], "total": 0,
        })
        slot["total"] += 1
        if e["verdict"] == LI_CANDIDATE_VERDICT:
            slot["candidates"].append(e)
        elif e["verdict"] == LI_VERIFY_VERDICT:
            slot["to_verify"].append(e)
        if e["flagged_but_kept"]:
            slot["flagged_but_kept"].append(e)

    return {
        "line_items": results,
        "by_module": by_module,
        "unresolved_view_names": unresolved_view_names,
        "summary": {
            "total_line_items_checked": len(results),
            "candidates_for_review": sum(1 for e in results if e["verdict"] == LI_CANDIDATE_VERDICT),
            "to_verify_conditional_formatting": sum(1 for e in results if e["verdict"] == LI_VERIFY_VERDICT),
            "modules_with_candidates": sum(1 for m in by_module.values() if m["candidates"]),
            "modules_with_verify_items": sum(1 for m in by_module.values() if m["to_verify"]),
            "flagged_but_kept": sum(1 for e in results if e["flagged_but_kept"]),
            "unresolved_view_names": len(unresolved_view_names),
        },
    }


def analyze(excel_path: Path, model_dir: Path) -> tuple:
    modules = load_modules_csv(model_dir)
    if not modules:
        raise SystemExit(f"No modules found in {model_dir / 'Modules.csv'} - check the path.")

    imports_usage = load_imports_csv(model_dir)
    ux_counts, excel_action_usage, line_item_exposure, ui_filter_modules = load_excel(excel_path)
    used_in_actions_all = imports_usage | excel_action_usage

    results = []
    for name, meta in modules.items():
        ux_count = ux_counts.get(name)
        entry = {
            "name": name,
            "functional_area": meta["functional_area"],
            "ux_count": ux_count,
            "referenced_by": meta["referenced_by"],
            "used_as_ui_filter": name in ui_filter_modules,
            "used_in_actions": name in used_in_actions_all,
            "is_category_header": meta["is_category_header"],
            "manual_marker": meta["manual_marker"],
        }

        # Category-header rows carry no real NUX/formula/import data by
        # definition, so they're resolved before - not instead of - the
        # ordered checks below; the outcome is identical either way.
        if meta["is_category_header"]:
            entry["verdict"] = "KEEP - module category header (audit trail)"
        elif ux_count is None:
            entry["verdict"] = "DATA MISMATCH - not found in NUX report; verify the module name matches exactly"
        elif ux_count > 0:
            entry["verdict"] = "ACTIVE - used in NUX pages/boards"
        elif meta["referenced_by"]:
            entry["verdict"] = "KEEP - feeds other modules via formula"
        elif entry["used_as_ui_filter"]:
            entry["verdict"] = "KEEP - used as a UI filter source on NUX pages"
        elif entry["used_in_actions"]:
            entry["verdict"] = "KEEP - used as an import/export source or target"
        else:
            entry["verdict"] = "CANDIDATE FOR REVIEW - no NUX usage, no formula reference, no UI-filter usage, no import/export usage found"

        entry["flagged_but_kept"] = meta["manual_marker"]["flagged"] and not entry["verdict"].startswith("CANDIDATE") and not entry["verdict"].startswith("DATA MISMATCH")

        results.append(entry)

    results.sort(key=lambda e: (
        not e["verdict"].startswith("CANDIDATE"),
        not e["manual_marker"]["flagged"],
        e["functional_area"],
        e["name"],
    ))
    report = {
        "modules": results,
        "summary": {
            "total_modules": len(results),
            "candidates_for_review": sum(1 for e in results if e["verdict"].startswith("CANDIDATE")),
            "active_in_ux": sum(1 for e in results if e["verdict"].startswith("ACTIVE")),
            "kept_internal_dependency": sum(1 for e in results if e["verdict"].startswith("KEEP")),
            "data_mismatches": sum(1 for e in results if e["verdict"].startswith("DATA MISMATCH")),
            "flagged_but_kept": sum(1 for e in results if e["flagged_but_kept"]),
        },
    }
    return report, line_item_exposure


def to_markdown_modules(report: dict, model_name: str) -> str:
    s = report["summary"]
    lines = [
        f"# Module optimization analysis - {model_name}",
        "",
        f"- Total modules: {s['total_modules']}",
        f"- Active in NUX: {s['active_in_ux']}",
        f"- Kept (internal dependency - formula/UI-filter/action usage, or category header): {s['kept_internal_dependency']}",
        f"- **Candidates for review: {s['candidates_for_review']}**",
        f"- Data mismatches (name not found in NUX report): {s['data_mismatches']}",
        f"- Flagged for deletion (Notes/Functional Area) but still active or kept: {s['flagged_but_kept']}",
        "",
    ]

    candidates = [e for e in report["modules"] if e["verdict"].startswith("CANDIDATE")]
    if candidates:
        lines += ["## Candidates for review", "", "| Module | Functional Area | Flagged for deletion |", "|---|---|---|"]
        for e in candidates:
            flag = ("Yes - " + "; ".join(e["manual_marker"]["reasons"])) if e["manual_marker"]["flagged"] else ""
            lines.append(f"| {e['name']} | {e['functional_area']} | {flag} |")
        lines.append("")

    mismatches = [e for e in report["modules"] if e["verdict"].startswith("DATA MISMATCH")]
    if mismatches:
        lines += ["## Data mismatches (need manual check)", "", "| Module | Functional Area |", "|---|---|"]
        for e in mismatches:
            lines.append(f"| {e['name']} | {e['functional_area']} |")
        lines.append("")

    kept = [e for e in report["modules"] if e["verdict"].startswith("KEEP")]
    if kept:
        lines += ["## Kept - internal dependency (zero NUX usage but still load-bearing)", "",
                   "| Module | Functional Area | Reason |", "|---|---|---|"]
        for e in kept:
            reason = e["verdict"].split(" - ", 1)[1]
            lines.append(f"| {e['name']} | {e['functional_area']} | {reason} |")
        lines.append("")

    flagged_but_kept = [e for e in report["modules"] if e["flagged_but_kept"]]
    if flagged_but_kept:
        lines += ["## Modules flagged for deletion but still active or kept", "",
                   "| Module | Functional Area | Verdict | Flag reason |", "|---|---|---|---|"]
        for e in flagged_but_kept:
            lines.append(f"| {e['name']} | {e['functional_area']} | {e['verdict']} | {'; '.join(e['manual_marker']['reasons'])} |")
        lines.append("")

    return "\n".join(lines)


def to_markdown_line_items(li_report: dict) -> str:
    s = li_report["summary"]
    lines = [
        "# Line item optimization analysis",
        "",
        f"- Total line items checked (modules kept/active at module level): {s['total_line_items_checked']}",
        f"- **Candidates for review: {s['candidates_for_review']}** across {s['modules_with_candidates']} module(s)",
        f"- **To verify (conditional formatting, not visible in NUX): {s['to_verify_conditional_formatting']}** across {s['modules_with_verify_items']} module(s)",
        f"- Flagged for deletion (Notes/inherited Functional Area) but still active or kept: {s['flagged_but_kept']}",
        f"- Unresolved view/module names in the NUX report: {s['unresolved_view_names']}",
        "",
        "## Line item candidates for review",
        "",
    ]

    any_candidates = False
    for module_name, data in sorted(li_report["by_module"].items()):
        if not data["candidates"]:
            continue
        any_candidates = True
        lines.append(f"### {module_name} ({data['functional_area']})")
        lines.append("")
        lines.append("| Line Item | Flagged for deletion |")
        lines.append("|---|---|")
        for e in data["candidates"]:
            flag = ("Yes - " + "; ".join(e["manual_marker"]["reasons"])) if e["manual_marker"]["flagged"] else (
                "Yes - inherited module Functional Area=DELETE" if e["inherited_module_delete_flag"] else "")
            lines.append(f"| {e['line_item']} | {flag} |")
        lines.append("")
    if not any_candidates:
        lines.append("None found.")
        lines.append("")

    lines.append("## Line items to verify usage (conditional formatting)")
    lines.append("")
    lines.append(
        "These names look like conditional-formatting drivers (`CF ...`) - Anaplan's format rules "
        "aren't captured anywhere in the NUX scrape, so a zero-usage signal here doesn't mean unused, "
        "only unconfirmed. Check the module's conditional formatting rules directly in Anaplan before "
        "acting on any of these."
    )
    lines.append("")
    any_verify = False
    for module_name, data in sorted(li_report["by_module"].items()):
        if not data["to_verify"]:
            continue
        any_verify = True
        lines.append(f"### {module_name} ({data['functional_area']})")
        lines.append("")
        lines.append("| Line Item | Flagged for deletion |")
        lines.append("|---|---|")
        for e in data["to_verify"]:
            flag = ("Yes - " + "; ".join(e["manual_marker"]["reasons"])) if e["manual_marker"]["flagged"] else (
                "Yes - inherited module Functional Area=DELETE" if e["inherited_module_delete_flag"] else "")
            lines.append(f"| {e['line_item']} | {flag} |")
        lines.append("")
    if not any_verify:
        lines.append("None found.")
        lines.append("")

    unresolved_view_names = li_report.get("unresolved_view_names", [])
    if unresolved_view_names:
        lines.append("## Data quality: unresolved view/module names")
        lines.append("")
        lines.append(
            "These names appeared in the NUX report's front-end exposure sheets but don't "
            "match any known module name in this model's CSV export - the view may have been "
            "renamed since the last scrape, so their exposure signal could not be matched to "
            "any line item. This is a data-quality flag, not a deletion signal."
        )
        lines.append("")
        for name in unresolved_view_names:
            lines.append(f"- {name}")
        lines.append("")

    flagged_but_kept = [e for m in li_report["by_module"].values() for e in m["flagged_but_kept"]]
    if flagged_but_kept:
        lines.append("## Line items flagged for deletion but still active or kept")
        lines.append("")
        lines.append("| Module | Line Item | Verdict | Flag reason |")
        lines.append("|---|---|---|---|")
        for e in flagged_but_kept:
            reasons = list(e["manual_marker"]["reasons"])
            if e["inherited_module_delete_flag"]:
                reasons.append("inherited module Functional Area=DELETE")
            lines.append(f"| {e['module']} | {e['line_item']} | {e['verdict']} | {'; '.join(reasons)} |")
        lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--model-dir", required=True, type=Path)
    parser.add_argument("--model-name", default=None)
    parser.add_argument("--out-json", type=Path, default=None)
    parser.add_argument("--out-markdown", type=Path, default=None)
    args = parser.parse_args()

    model_name = args.model_name or args.model_dir.name
    module_report, line_item_exposure = analyze(args.excel, args.model_dir)
    li_report = analyze_line_items(args.model_dir, module_report["modules"], line_item_exposure)

    md = to_markdown_modules(module_report, model_name) + "\n" + to_markdown_line_items(li_report)
    combined = {"modules": module_report, "line_items": li_report}

    if args.out_json:
        args.out_json.parent.mkdir(parents=True, exist_ok=True)
        args.out_json.write_text(json.dumps(combined, indent=2), encoding="utf-8")
    if args.out_markdown:
        args.out_markdown.parent.mkdir(parents=True, exist_ok=True)
        args.out_markdown.write_text(md, encoding="utf-8")

    try:
        print(md)
    except UnicodeEncodeError:
        print(md.encode(sys.stdout.encoding or "utf-8", errors="replace").decode(sys.stdout.encoding or "utf-8"))


if __name__ == "__main__":
    main()
